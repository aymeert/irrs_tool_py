from openpyxl import load_workbook
from openpyxl.styles import Font

path_mac_ar_in = "/Users/aymeerodriguez/Documents/GitHub/irrs_tool_py/example.xlsx"
path_mac_ar_out = "/Users/aymeerodriguez/Documents/GitHub/irrs_tool_py/example_changed.xlsx"
path_mac_jm = "/Users/javier/Documents/GitHub/irrs_tool_py/example.xlsx"
path_mac_jm2 = "/Users/javier/Documents/GitHub/irrs_tool_py/example_changed.xlsx"
path = "C:\\Users\\aymee.rodriguez\\OneDrive - Exactech, Inc\\Projects\\irrs_tool_py\\example.xlsx"
path2 = "C:\\Users\\aymee.rodriguez\\OneDrive - Exactech, Inc\\Projects\\irrs_tool_py\\example_changed.xlsx"
path_full_irss = "C:\\Users\\aymee.rodriguez\\OneDrive - Exactech, Inc\\Projects\\irrs_tool_py\\QC322-110-00 Rev A 2022-07-15-15-22-38.xlsx"
path_full_irss_mac = "/Users/javier/Documents/GitHub/irrs_tool_py/QC322-110-00 Rev A 2022-07-15-15-22-38.xlsx"
path_to_translation_table = "/Users/javier/Documents/GitHub/irrs_tool_py/translation_table.xlsx"

def open_workbook(path_to_workbook):
    """brief description
    
    Args:
        Object in
    
    Returns:
        Object out
    """
    workbook = load_workbook(path_to_workbook)
    worksheet = workbook["Final Or Supplier Manufactured"]
    return workbook, worksheet


def find_bp_specification(worksheet):
    """brief description
    
    Args:
        Object in
    
    Returns:
        Object out
    """
    start_cell = "BP Specification"
    for col in range(worksheet.min_column, worksheet.max_column):
        for row in range(worksheet.min_row, worksheet.max_row):
            if worksheet.cell(row,col).value == start_cell:
                start_row, start_col = row, col
                break
    return start_row, start_col


def iterate_through_column(worksheet):
    """brief description
    
    Args:
        Object in
    
    Returns:
        Object out
    """
    start_row, start_col = find_bp_specification(worksheet)
    for row in range(start_row + 1, worksheet.max_row):
        cell = worksheet.cell(row,start_col)
        if cell.value is None:
            break
        cell.font = Font(name= 'Y14.5-2009', size=12)
        translated_symbols = translate_by_cell_type(cell)
        cell.value = translated_symbols
    return worksheet


def frame_simple_cell(cell):
    """brief description
    
    Args:
        Object in
    
    Returns:
        Object out
    """
   # This should be a function
    wrong_symbols = str(cell.value)
    right_symbols = wrong_symbols.replace("|", "{", 1)
    right_symbols = right_symbols[::-1].replace("|", "}", 1)
    right_symbols = right_symbols[::-1]
    right_symbols = right_symbols.replace("⌖", "¿~", 1)
    right_symbols = right_symbols.replace("Ⓜ", "Ì~", 1)
    
    # This should be another function
    first_half = right_symbols.split('{')[0]
    second_half = right_symbols.split('{')[1]
    alphabet = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz."
    final_symbols = ""
    for letter in second_half:
        if letter in alphabet:
            letter = letter + "`"
        final_symbols = final_symbols + letter
    final_symbols = first_half+'{'+final_symbols
    return final_symbols


def read_translation_table(path_to_translation_table):
    workbook = load_workbook(path_to_translation_table)
    translation_table = workbook["Translations"]
    return translation_table


def translate_gdt_symbols(cell):
    #str(cell.value)
    right_symbols = "2X |⌖|.009Ⓜ|B|"
    translation_table = read_translation_table(path_to_translation_table)
    for row in range(translation_table.min_row + 1, translation_table.max_row):
        correct_symbol = str(translation_table.cell(row,2).value)
        incorrect_symbol = translation_table.cell(row,3).value #not reading it as a string initially
        if incorrect_symbol:
            right_symbols = right_symbols.replace(str(incorrect_symbol), correct_symbol + "~", 1)
    print(right_symbols)
    return


def translate_by_cell_type(cell):
    """brief description
    
    Args:
        Object in
    
    Returns:
        Object out
    """
    cell_content = cell.value
    if is_simple_frame(cell_content):
        translated_cell = frame_simple_cell(cell)
    else: translated_cell = cell_content
    return translated_cell

def is_simple_frame(cell_content):
    """identifies if a cell's content is a simple GDT frame
    
    Args:
        String: the contents of a cell wiht a gdt tag to be identified
    
    Returns:
        Boolean: wether the string is a simple frame or not
    """
    if cell_content.count("|") >= 2:
        return True
    else: return False

workbook, worksheet  = open_workbook(path_full_irss_mac)
translated_worksheet = iterate_through_column(worksheet)
#testing
translate_gdt_symbols(worksheet)

# workbook.save(path_mac_jm2)