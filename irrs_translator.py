from openpyxl import load_workbook
from openpyxl.styles import Font
import openpyxl.cell._writer
from pathlib import Path
import tkinter as tk
from tkinterdnd2 import DND_FILES, TkinterDnD

path_to_translation_table = Path("J:\\Public\\Employee\\AYMEE.RODRIGUEZ\\IRRS translator program\\translation_table.xlsx")
path_to_template = Path("J:\\Public\\Employee\\AYMEE.RODRIGUEZ\\IRRS translator program\\701-104-004.xlsx")

def open_workbook(path_to_workbook):
    """Opens the IRRS to be translated"""
    workbook = load_workbook(path_to_workbook)
    worksheet = workbook["Final Or Supplier Manufactured"]
    return workbook, worksheet


def find_bp_specification(worksheet):
    """Finds the column header to be translated"""
    start_cell = "BP Specification"
    for col in range(worksheet.min_column, worksheet.max_column):
        for row in range(worksheet.min_row, worksheet.max_row):
            if worksheet.cell(row,col).value == start_cell:
                start_row, start_col = row, col
                break
    return start_row, start_col


def find_word_in_worksheet(worksheet, word):
    """Finds the coordinates of a cell"""
    cell_row, cell_col = -1, -1
    for col in range(worksheet.min_column, worksheet.max_column + 1):
        for row in range(worksheet.min_row, worksheet.max_row + 1):
            # print(str(worksheet.cell(row,col).value))
            # print(word)
            if ((str(worksheet.cell(row,col).value)).find(word) != -1):
                cell_row, cell_col = row, col
                break
    return cell_row, cell_col


def iterate_through_column(worksheet):
    """Accesses cell by cell in the column to be translated"""
    start_row, start_col = find_bp_specification(worksheet)
    for row in range(start_row + 1, worksheet.max_row + 1):
        cell = worksheet.cell(row,start_col)
        if cell.value is None:
            break
        cell.font = Font(color='FF000000', name= 'Y14.5-2009', size=13)
        translated_symbols = translate_by_cell_type(cell)
        cell.value = translated_symbols
    return worksheet


def translate_by_cell_type(cell):
    """Identifies the kind of cell to use the appropriate translation method"""
    cell_content = cell.value
    if is_composite_frame(cell_content):
        translated_cell = frame_composite_cell(cell_content)
    elif is_surface_finish(cell_content):
        translated_cell = fix_surface_finish(cell_content)
    elif is_stacked_frame(cell_content):
        translated_cell = frame_stacked_cell(cell_content)
    elif is_simple_frame(cell_content):
        translated_cell = frame_simple_cell(cell_content)
    else: translated_cell = translate_gdt_symbols(cell_content) # this was changed from cell_content
    return translated_cell


def is_simple_frame(cell_content):
    """Identifies if a cell's content is a simple GD&T frame"""
    if cell_content.count("|") >= 2 and cell_content.count("FRAME") == 0 and cell_content.count("<&80>") == 0:
        return True
    else: return False


def is_composite_frame(cell_content):
    """Identifies if a cell's content is a composite frame"""
    if cell_content.count("FRAME") == 2:
        return True
    else: return False 


def is_stacked_frame(cell_content):
    """Identifies if a cell's content is a stacked frame""" 
    if cell_content.count("<&80>") == 1:
        return True
    else: return False


def is_surface_finish(cell_content):
    """Identifies if a cell's content is surface finish"""
    if cell_content.count("(|Lay Symbol:R<L>a<L>") == 1:
        return True
    else: return False


def frame_simple_cell(cell_content):
    """Adds framing to cells that require a single GD&T frame"""
    translated_symbols = translate_gdt_symbols(cell_content)
    translated_symbols = translated_symbols.replace("|", "{", 1)
    translated_symbols = translated_symbols[::-1].replace("|", "}", 1)
    translated_symbols = translated_symbols[::-1]
    before_first_braket = translated_symbols.split('{')[0]
    after_first_braket = translated_symbols.split('{')[1]
    before_second_braket = after_first_braket.split('}')[0]
    after_second_braket = after_first_braket.split('}')[1]
    framed_characters = add_frames_to_characters(before_second_braket)
    complete_translation = before_first_braket + '{' + framed_characters + '}' + after_second_braket
    return complete_translation


def frame_composite_cell(cell_content):
    """Adds framing to cells that require a composite GD&T cell"""
    composite_frame = cell_content.replace("FRAME 1|","",1)
    before_frame_2 = composite_frame.split("FRAME 2")[0]
    after_frame_2 = composite_frame.split("FRAME 2")[1]
    translated_frame_1 = frame_simple_cell(before_frame_2)
    frame_1_symbol = translated_frame_1.split("|")[0]
    translated_frame_2 = frame_simple_cell(after_frame_2)
    translated_frame_2 = frame_1_symbol + translated_frame_2.replace( "{","|",1)
    composite_frame = translated_frame_1 + "  " + translated_frame_2
    return composite_frame 


def frame_stacked_cell(cell_content): #This only works for 2 stacked lines, needs improvement for more than 2
    """Adds framing to cells that require a stacked GD&T cell"""
    stacked_frame = cell_content.replace("<&80>","s",1)
    before_symbol = stacked_frame.split("s")[0]
    after_symbol = stacked_frame.split("s")[1]
    translated_frame_1 = frame_simple_cell(before_symbol)
    translated_frame_2 = frame_simple_cell(after_symbol)
    stacked_frame = translated_frame_1 + "  " + translated_frame_2
    return stacked_frame


def fix_surface_finish(cell_content):
    """Fixes surface finish output from NX generator report"""
    surface_finish = cell_content.replace("(|Lay Symbol:R<L>a<L>"," Ra",1)
    surface_finish = surface_finish.replace(")","",1)
    return surface_finish


def translate_gdt_symbols(cell_content):
    """Translates the GD&T symbols from the translation table"""
    translated_symbols = cell_content # changing it from this str(cell.value) to cell_content
    translation_table = read_translation_table(path_to_translation_table)
    translated_symbols = translated_symbols.replace("<o>","Ø") #handling special case
    translated_symbols = translated_symbols.replace(",","") #handling special case
    for row in range(translation_table.min_row + 1, translation_table.max_row):
        correct_symbol = str(translation_table.cell(row,2).value)
        incorrect_symbol = translation_table.cell(row,3).value #not reading it as a string initially
        if incorrect_symbol:
            translated_symbols = translated_symbols.replace(str(incorrect_symbol), correct_symbol)
    return translated_symbols


def read_translation_table(path_to_translation_table):
    """Reads the translation table containing the GD&T symbols"""
    workbook = load_workbook(path_to_translation_table)
    translation_table = workbook["Translations"]
    return translation_table


def add_frames_to_characters(characters_to_frame):
    """Adds frames to individual alpha numeric characters and special GD&T symbols"""
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz"
    numbers = "0123456789"
    punctuation = "."
    gdt_symbols = get_list_of_gdt_symbols(path_to_translation_table)
    framed_characters = ""
    for character in characters_to_frame:
        if character in alphabet:
            character = character + "_"
        elif character in numbers:
            character = character + "`"
        elif character in punctuation:
            character = character + "\\"
        elif character in gdt_symbols:
            character = character + "~"
        framed_characters = framed_characters + character
    return framed_characters


def get_list_of_gdt_symbols(path_to_translation_table):
    """Gets a list of all the GD&T symbols from the translation table"""
    translation_table = read_translation_table(path_to_translation_table) # change the path depending on the OS
    gdt_symbols = ""
    for row in range(translation_table.min_row + 1, translation_table.max_row):
        incorrect_symbol = translation_table.cell(row,3).value #not reading it as a string initially
        if incorrect_symbol:
            gdt_symbols = gdt_symbols + str(translation_table.cell(row,2).value)
    return gdt_symbols


def generate_irrs_output_path(path_to_irrs_for_translation):
    """Generate the output path for the translated IRRS with the same name"""
    path_to_translated_irrs = path_to_irrs_for_translation.replace(".xlsx", " [Translated].xlsx")
    path_to_translated_irrs = Path(path_to_translated_irrs)
    return path_to_translated_irrs


def copy_to_correct_template(worksheet_irrs, worksheet_template):
    """Copies IRRS contents to the correct template"""
    start_row_irrs, start_col_irrs = find_bp_specification(worksheet_irrs)
    start_row_template, start_col_template = find_bp_specification(worksheet_template)
    upper_row, upper_col = find_word_in_worksheet(worksheet_irrs, "Upper Limit") #change back to "Upper Limit" to only copy "IRRS Comments"
    irrs_comment_row, irrs_comment_col = find_word_in_worksheet(worksheet_irrs, "IRRS Comments")


# Changing color to black
    worksheet_template.cell(row = 5, column = 4).font = Font(color='FF000000')
    worksheet_template.cell(row = 5, column = 13).font = Font(color='FF000000')
    worksheet_template.cell(row = 4, column = 15).font = Font(color='FF000000')
    worksheet_template.cell(row = 6, column = 13).font = Font(color='FF000000')
    worksheet_template.cell(row = 6, column = 4).font = Font(color='FF000000')
    worksheet_template.cell(row = 4, column = 4).font = Font(color='FF000000')

    relevant_cells = ["Sampling Plan (IRRS) Name:","Product Number:","No. of Inspection Lines:"]
    # Unmerging cell ranges for the relevant cells
    sampling_plan_name_nx = "E5:G5"
    sampling_plan_name_temp = "D5:G5"
    worksheet_irrs.unmerge_cells(range_string = sampling_plan_name_nx)
    worksheet_template.unmerge_cells(range_string = sampling_plan_name_temp)

    sampling_plan_rev_nx = "E6:G6" #new
    sampling_plan_rev_temp = "D6:G6"
    worksheet_irrs.unmerge_cells(range_string = sampling_plan_rev_nx)
    worksheet_template.unmerge_cells(range_string = sampling_plan_rev_temp)
    
    number_inspection_lines_nx = "O4:S4"
    number_inspection_lines_temp = "O4:S4"
    worksheet_irrs.unmerge_cells(range_string = number_inspection_lines_nx)
    worksheet_template.unmerge_cells(range_string = number_inspection_lines_temp)

    DCRTC_nx = "E4:M4" #new
    DCRTC_temp = "D4:M4"
    worksheet_irrs.unmerge_cells(range_string = DCRTC_nx)
    worksheet_template.unmerge_cells(range_string = DCRTC_temp)

    # Copying specific content
    worksheet_template.cell(row = 5, column = 4).value = worksheet_irrs.cell(row = 5, column = 5).value #sampling plan name
    worksheet_template.cell(row = 5, column = 13).value = worksheet_irrs.cell(row = 5, column = 13).value #Product number
    worksheet_template.cell(row = 4, column = 15).value = worksheet_irrs.cell(row = 4, column = 15).value #No. of inspection lines
    worksheet_template.cell(row = 6, column = 13).value = worksheet_irrs.cell(row = 6, column = 13).value # product revision
    worksheet_template.cell(row = 6, column = 4).value = worksheet_irrs.cell(row = 6, column = 5).value #sampling plan revision
    worksheet_template.cell(row = 4, column = 4).value = worksheet_irrs.cell(row = 4, column = 5).value #DCRTC
    
    # Merging the cells
    worksheet_irrs.merge_cells(range_string = sampling_plan_name_nx)
    worksheet_template.merge_cells(range_string = sampling_plan_name_temp)

    worksheet_irrs.merge_cells(range_string = sampling_plan_rev_nx) #new
    worksheet_template.merge_cells(range_string = sampling_plan_rev_temp)
    
    worksheet_irrs.merge_cells(range_string = number_inspection_lines_nx)
    worksheet_template.merge_cells(range_string = number_inspection_lines_temp)

    worksheet_irrs.merge_cells(range_string = DCRTC_nx)     #new
    worksheet_template.merge_cells(range_string = DCRTC_temp)

    
    # Clearing the template
    for row in range(start_row_irrs + 1, worksheet_template.max_row + 1):
        for col in range(start_col_irrs - 1, irrs_comment_col + 1):
            worksheet_template.cell(row = row, column = col).value = None
            worksheet_template.cell(row = row, column = col).font = Font(color='FF000000')

    # Copying from IRRS from NX to IRRS Template
    for row in range(start_row_irrs + 1, worksheet_irrs.max_row + 1):
        for col in range(start_col_irrs - 1, upper_col + 1):
            worksheet_template.cell(row = row, column = col).value = worksheet_irrs.cell(row = row, column = col).value
    return worksheet_template


def translate_irrs_button_logic():
    """Main logic for window GUI"""
    files_translated_counter = 0
    workbook_template, worksheet_template = open_workbook(path_to_template) # opening the correct template
    for element_index, list_element in enumerate(list_irrs_path.get(0,tk.END)):
        files_translated_counter += 1
        if list_element:
            path_check_if_valid_irrs = Path(list_element)
            if path_check_if_valid_irrs.is_file():
                path_to_translated_irrs = generate_irrs_output_path(list_element)
                if not path_to_translation_table.is_file():
                    label_footer["text"] = "Translation Table not available. Check VPN connection. For help, email: aymee.rodriguez@exac.com"
                # try: 
                workbook_irrs, worksheet_irrs = open_workbook(Path(list_element))
                worksheet_template = copy_to_correct_template(worksheet_irrs, worksheet_template)
                worksheet_template = iterate_through_column(worksheet_template) #turn me back on
                workbook_template.save(path_to_translated_irrs)
                workbook_template.close()
                # except:
                #     files_translated_counter -= 1
                #     continue
                if len(list_irrs_path.get(0,tk.END)) == files_translated_counter:
                    label_footer["text"] = "Translation successful! " + str(files_translated_counter) + " file(s) saved in the same location with [Translated] appended"
            else:
                label_footer["text"] = "Path provided is not valid, try again"
        else:
            label_footer["text"] = "Input is not valid, try again"
    
    list_irrs_path.delete(0,tk.END)


def read_tabulated_data(path_to_tabulated):
    """Opens the tabulated spreadsheet"""
    workbook_tabulated = load_workbook(path_to_tabulated)
    worksheet_tabulated = workbook_tabulated.active
    return workbook_tabulated, worksheet_tabulated

mandatory_columns = ["DCRTC#", "Sampling Plan (IRRS) Name", "Product Description", "Product Number", "Sampling Plan Revision", "Product Revision"]

def add_columns_to_tabulated_table(workbook_tabulated, worksheet_tabulated):
    """Adds any missing mandatory columns to the tabulated data file"""
    column_names = []
    for col in range(1, worksheet_tabulated.max_column + 1):
        column_names.append(worksheet_tabulated.cell(1, column = col).value)
    columns_to_be_added = [element for element in mandatory_columns if element not in column_names]
    counter = 0
    for col in range(worksheet_tabulated.max_column + 1, worksheet_tabulated.max_column + len(columns_to_be_added) + 1):
        worksheet_tabulated.cell(1, column = col).value = columns_to_be_added[counter]
        print(columns_to_be_added[counter])
        counter += 1
    return worksheet_tabulated


def is_tabulated_data(path_to_file):
    """Cheks that the file is not an IRRS"""
    try:
        open_workbook(path_to_file)
    except KeyError:
        return True
    else:
        return False


def generate_tabulated_irrs():
    """Generates tabulated IRRSs"""
    for element_index, list_element in enumerate(list_irrs_path.get(0,tk.END)):
        if (list_element.upper()).find("TABULATED DATA") != -1:
            workbook_tabulated, worksheet_tabulated = read_tabulated_data(Path(list_element))
            # write logic to check correct columns exist and that data has been added
        else:
            path_to_irrs_for_tabulation = list_element
            workbook_irrs, worksheet_irrs = open_workbook(Path(path_to_irrs_for_tabulation))
    
        # worksheet_tabulated = add_columns_to_tabulated_table(workbook_tabulated, worksheet_tabulated)
        # workbook_tabulated.save(Path(list_element))
        # if is_tabulated_data(Path(list_element)):
        #     button_generate_irrs["text"] = "Add Columns to Table"
    # getting column names
    column_names = []
    for col in range(1, worksheet_tabulated.max_column + 1):
        if (worksheet_tabulated.cell(1, column = col).value == "Sampling Plan (IRRS) Name") :
            irrs_name_column_no = col
        if (worksheet_tabulated.cell(1, column = col).value in mandatory_columns) or ((worksheet_tabulated.cell(1, column = col).value).find("DIM") != -1) :
            column_names.append(worksheet_tabulated.cell(1, column = col).value) #checking which columns in the tabulated data table are mandatory or have the word DIM
            # because only those need to be tabulated
    # if I have column names then I can checka against the ones that should be translated, then have a an if statement that copies over the content
    # acording to the cell type, for instance some need unmerging while ithers don't
    
    # define a loop to make many copies as there are rows in the tabulated table then
    # define a loop to loop through all the DIMs
    # check that there are as nany DIMs in the tabulated data as in the IRRs

    #create an empty dictionary to add IRRS names and paths:
    irrs_to_tabulate = {}

    for dim_range in range(2, worksheet_tabulated.max_row + 1):
        workbook_irrs, worksheet_irrs = open_workbook(Path(path_to_irrs_for_tabulation))
        workbook_processing = workbook_irrs
        worksheet_processing = worksheet_irrs
        # the fisrt copy of an IRRS happens here
        for dim_number in range(1, worksheet_tabulated.max_column + 1):
            # this statement is kept at row 1 because this will be the projection of the headers
            irrs_filename = worksheet_tabulated.cell(row = dim_range, column = irrs_name_column_no).value #this is getting the name that the tabulated IRRS should be saved as
        ###########################################################################################################################################
        # for dim_number in range(1, worksheet_tabulated.max_column + 1):
        #     # this statement is kept at row 1 because this will be the projection of the headers
        #     irrs_filename = worksheet_tabulated.cell(row = dim_range, column = irrs_name_column_no).value #this is getting the name that the tabulated IRRS should be saved as
        #     if (worksheet_tabulated.cell(row = 1, column = dim_number).value in column_names):
        #         header_value = str(worksheet_tabulated.cell(row = 1, column = dim_number).value)
        #         cell_value = str(worksheet_tabulated.cell(row = dim_range, column = dim_number).value)
        #         if (header_value.find("DIM") != -1):
        #             print(dim_range, dim_number)
        #             row_v, col_v = find_word_in_worksheet(worksheet_processing, header_value)
        #             cell_to_process = worksheet_processing.cell(row = row_v, column = col_v).value
        #             worksheet_processing.cell(row = row_v, column = col_v).value = cell_to_process.replace(header_value, cell_value)
        ###########################################################################################################################################
        path_to_irrs_for_tabulation = make_tabulated_outpute_name(path_to_irrs_for_tabulation, irrs_filename)
        # print(path_to_irrs_for_tabulation)
        #assigning values to dictionary
        irrs_to_tabulate[dim_range] = path_to_irrs_for_tabulation

        workbook_processing.save(path_to_irrs_for_tabulation)
        workbook_processing.close()
        # we need to repurpose the loop we had to repalce the DIMs such that the inputs are now the new IRRS files we created
        # we may need to create a dictionay with the names and the paths and open them in order and then tabulate them in the same order as the loops
                    
        # now the file is saved with a number appended at the end
    for key, irrs in irrs_to_tabulate.items():
        for dim_range in range(key, key + 1):
            workbook_tabulating, worksheet_tabulating = open_workbook(irrs)
            for dim_number in range(1, worksheet_tabulated.max_column + 1):
            # this statement is kept at row 1 because this will be the projection of the headers
                irrs_filename = worksheet_tabulated.cell(row = dim_range, column = irrs_name_column_no).value #this is getting the name that the tabulated IRRS should be saved as
                if (worksheet_tabulated.cell(row = 1, column = dim_number).value in column_names):
                    header_value = str(worksheet_tabulated.cell(row = 1, column = dim_number).value)
                    cell_value = str(worksheet_tabulated.cell(row = dim_range, column = dim_number).value)
                    if (header_value.find("DIM") != -1):
                        print(dim_range, dim_number)
                        row_v, col_v = find_word_in_worksheet(worksheet_tabulating, header_value)
                        cell_to_process = worksheet_tabulating.cell(row = row_v, column = col_v).value
                        worksheet_tabulating.cell(row = row_v, column = col_v).value = cell_to_process.replace(header_value, cell_value)
            workbook_tabulating.save(irrs)
            workbook_tabulating.close()
    print(irrs_to_tabulate)
    return


def make_tabulated_outpute_name(path_to_irrs_for_tabulation, irrs_filename):
    """Generate the output path for the tabulated IRRS with the same name and dimension"""
    tabulated_irrs_filename = str(irrs_filename) + ".xlsx"
    path_to_tabulated_irrs = Path(path_to_irrs_for_tabulation)
    parent_path = path_to_tabulated_irrs.parent
    path_to_irrs_for_tabulation = Path(str(parent_path) + "\\" + tabulated_irrs_filename)
    return path_to_irrs_for_tabulation

def copy_tabulated_to_irrs():
    return worksheet_processed

def add_paths_to_listbox(event):
    """Adds file paths to listbox and checks that no duplicates are added"""
    paths_list = event.data
    paths_list = paths_list.split('} {')
    items_in_listbox = {list_element for element_index, list_element in enumerate(list_irrs_path.get(0,tk.END)) if list_element}
    paths_list = [path.replace('{','') for path in paths_list]
    paths_list = [path.replace('}','') for path in paths_list]
    set_of_paths = set(paths_list)
    set_of_paths = set_of_paths.difference(items_in_listbox) # only adding unique items
    for path in set_of_paths:
        list_irrs_path.insert("end",path)
    

def clear_list_button_logic():
    """Clears the list if mistakes were made"""
    list_irrs_path.delete(0, tk.END)
    label_footer["text"] = "List cleared, add files for translation"





window = TkinterDnD.Tk()  # notice - use this instead of tk.Tk()

window.title("Exactech IRRS Translator v 0.30")
icon_path = Path("J:\\Public\\Employee\\AYMEE.RODRIGUEZ\\IRRS translator program\\exactech.ico")
if icon_path.is_file():
    window.iconbitmap(icon_path)
window.resizable(False, False)
window.columnconfigure(0, minsize = 250)
window.rowconfigure([0, 4], minsize = 50) 

label_title = tk.Label( text = "Drag IRRS to be translated here: ")
label_title.grid(row = 0, column = 0)

scrollbar_x = tk.Scrollbar(orient = "horizontal")
scrollbar_y = tk.Scrollbar(orient = "vertical")
list_irrs_path = tk.Listbox(width = 100, xscrollcommand = scrollbar_x.set, yscrollcommand = scrollbar_y.set) 
list_irrs_path.grid(row = 1, column = 0)
list_irrs_path.drop_target_register(DND_FILES)
list_irrs_path.dnd_bind('<<Drop>>', add_paths_to_listbox) #lambda e: list_irrs_path.insert(tk.END, e.data)

scrollbar_x.config(command = list_irrs_path.xview)
scrollbar_x.grid(row = 2, column = 0, sticky = 'ew')
scrollbar_y.config(command = list_irrs_path.xview)
scrollbar_y.grid(row = 1, column = 1, sticky = 'ns')

button_translate_irrs = tk.Button(text = "Translate", command = translate_irrs_button_logic)
button_translate_irrs.grid(row = 3, column = 0, pady = 10) 

button_generate_irrs = tk.Button(text = "Generate Tabulated IRRS", command = generate_tabulated_irrs)
button_generate_irrs.grid(row = 4, column = 0) 

button_clear = tk.Button(text="Clear List", command = clear_list_button_logic)
button_clear.grid(row = 5, column = 0) 

label_footer = tk.Label(text = "")
label_footer.grid(row = 6, column = 0, pady = 10)

# to build:
# open terminal and navigate to C:\Users\aymee.rodriguez\irrs\Scripts by typing cd 
# run: activate.bat - (irrs) should appear before working directory
# replace irrs_translator.py with new one from VScode
# the run: pyinstaller -F -w --icon=exactech.ico irrs_translator.py --additional-hooks-dir=.
# open C:\Users\aymee.rodriguez\irrs\Scripts\dist and copy the irrs_translator.exe to the shared folder
window.mainloop()

# TODO:
"""
    [X] add a function to read an excel table with the codes
    translation
    [X] modify the processIRRS function to replace the symbols based
    on the translation table
    [X] add a function to read each IRRS to be translated
    [X] add a function to export each translated IRRS
    [] replace all string concatenation with ''.join() #performance
    [X] add a function to create the exported file with the same name as orignal and in the same folder
    [X] add a function to get the active directory and use the translation table in that directory
    [X] add 5 cases:
        [X] simple frame
            [X] simple frame with text at the end
        [X] double frame (there are two double frame symbols; not sure if NX output is different for each) 
           [X] composite (we have this case)
           [X] stacked 
        [X] not framed but translated
        [X] Ra (check symbols in NX)
        [X] nothing needs to happen
    [X] create a graphical (or terminal) user interface
    [X] create an executable program to distribute
    [X] remove {} when at the begining and the end of a string path
    [X] reduce size of the program to increase speed #performance
    [] use list comprhension when available #performance
    [X] Improve terminal user interface
    [X] Create survey to see what the users think 
    [X] Delete path after translation is over
    [X] Bigger box 
    [X] Able to add several IRRS to be translated at the same time
    [X] Make a list for the drag and drop box so that it is easier for the user to see
    [X] Add program icon
    [X] Fix vertical bar
    [X] Find best size for the window
    [X] Add multiple files at once instead of dropping one by one
    [X] check if translation table exists, if not ask for new directory
    [] only let the user run the program if they are using the latest version
    [X] Add buttom to clear entry box
    [] Clear second sheet
"""
